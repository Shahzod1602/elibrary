from django.contrib import admin, messages
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import path, reverse
from django.utils import timezone

from openpyxl import load_workbook

from .models import Book, Category, Review, Favorite, Exhibit, Event, News, BorrowRecord


@admin.register(Category)
class CategoryAdmin(admin.ModelAdmin):
    list_display = ['name', 'icon']
    search_fields = ['name']


@admin.register(Book)
class BookAdmin(admin.ModelAdmin):
    list_display = [
        'title',
        'author',
        'category',
        'isbn',
        'barcode',
        'quantity',
        'language',
        'available',
        'views_count',
        'created_at',
    ]
    list_filter = ['category', 'language', 'available', 'published_year']
    search_fields = ['title', 'author', 'isbn', 'barcode']
    list_editable = ['available', 'quantity']
    readonly_fields = ['views_count', 'download_count']
    change_list_template = 'admin/books/book/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'circulation/',
                self.admin_site.admin_view(self.circulation_view),
                name='books_book_circulation',
            ),
            path(
                'import-xlsx/',
                self.admin_site.admin_view(self.import_xlsx_view),
                name='books_book_import_xlsx',
            ),
        ]
        return custom_urls + urls

    @staticmethod
    def _normalize_header(value):
        if value is None:
            return ''
        return str(value).strip().lower().replace(' ', '').replace('_', '')

    @staticmethod
    def _to_text(value):
        if value is None:
            return ''
        text = str(value).strip()
        if text.endswith('.0') and text.replace('.', '', 1).isdigit():
            return text[:-2]
        return text

    @staticmethod
    def _to_int(value, default=1):
        if value is None or str(value).strip() == '':
            return default
        try:
            parsed = int(float(str(value).strip()))
            return max(parsed, 0)
        except (TypeError, ValueError):
            return default

    @staticmethod
    def _resolve_student(identifier):
        value = (identifier or '').strip()
        if not value:
            return None, 'Student identifier is required.'

        matches = User.objects.filter(
            Q(username__iexact=value) |
            Q(email__iexact=value) |
            Q(student_profile__id_number__iexact=value)
        ).distinct()

        if matches.count() == 0:
            return None, 'Student not found. Use ID number, username, or email.'
        if matches.count() > 1:
            return None, 'Multiple students matched this identifier. Use a unique ID number.'
        return matches.first(), None

    def circulation_view(self, request):
        if request.method == 'POST':
            action = request.POST.get('action_type')
            barcode = (request.POST.get('barcode') or '').strip()
            student_identifier = (request.POST.get('student_identifier') or '').strip()

            if not barcode:
                self.message_user(request, 'Barcode is required.', level=messages.ERROR)
                return HttpResponseRedirect(request.path)

            if action == 'checkout':
                student, error = self._resolve_student(student_identifier)
                if error:
                    self.message_user(request, error, level=messages.ERROR)
                    return HttpResponseRedirect(request.path)

                try:
                    with transaction.atomic():
                        book = Book.objects.select_for_update().get(barcode=barcode)
                        if book.quantity < 1:
                            self.message_user(
                                request,
                                f'"{book.title}" is out of stock (quantity=0).',
                                level=messages.ERROR,
                            )
                            return HttpResponseRedirect(request.path)

                        BorrowRecord.objects.create(
                            book=book,
                            student=student,
                            assigned_by=request.user,
                        )
                        book.quantity -= 1
                        book.available = book.quantity > 0
                        book.save(update_fields=['quantity', 'available'])
                except Book.DoesNotExist:
                    self.message_user(request, 'Book not found for this barcode.', level=messages.ERROR)
                    return HttpResponseRedirect(request.path)

                self.message_user(
                    request,
                    f'Checked out: "{book.title}" -> {student.username}. Remaining quantity: {book.quantity}.',
                    level=messages.SUCCESS,
                )
                return HttpResponseRedirect(request.path)

            if action == 'checkin':
                active_records = BorrowRecord.objects.select_related('book', 'student').filter(
                    book__barcode=barcode,
                    checked_in_at__isnull=True,
                )

                if student_identifier:
                    student, error = self._resolve_student(student_identifier)
                    if error:
                        self.message_user(request, error, level=messages.ERROR)
                        return HttpResponseRedirect(request.path)
                    active_records = active_records.filter(student=student)

                if not active_records.exists():
                    self.message_user(
                        request,
                        'No active checkout found for this barcode.',
                        level=messages.ERROR,
                    )
                    return HttpResponseRedirect(request.path)

                if active_records.count() > 1 and not student_identifier:
                    self.message_user(
                        request,
                        'Multiple students have this barcode checked out. Enter student ID/username for check-in.',
                        level=messages.ERROR,
                    )
                    return HttpResponseRedirect(request.path)

                record = active_records.order_by('-checked_out_at').first()

                with transaction.atomic():
                    locked_book = Book.objects.select_for_update().get(pk=record.book_id)
                    record.checked_in_at = timezone.now()
                    record.returned_by = request.user
                    record.save(update_fields=['checked_in_at', 'returned_by'])

                    locked_book.quantity += 1
                    locked_book.available = locked_book.quantity > 0
                    locked_book.save(update_fields=['quantity', 'available'])

                self.message_user(
                    request,
                    f'Checked in: "{record.book.title}" from {record.student.username}. Quantity: {locked_book.quantity}.',
                    level=messages.SUCCESS,
                )
                return HttpResponseRedirect(request.path)

            self.message_user(request, 'Unknown action.', level=messages.ERROR)
            return HttpResponseRedirect(request.path)

        context = dict(
            self.admin_site.each_context(request),
            opts=self.model._meta,
            title='Book circulation (Check In / Check Out)',
            active_records=BorrowRecord.objects.select_related('book', 'student')[:20],
        )
        return render(request, 'admin/books/book/circulation.html', context)

    def import_xlsx_view(self, request):
        if request.method == 'POST':
            file_obj = request.FILES.get('xlsx_file')
            if not file_obj:
                self.message_user(request, 'Please select an .xlsx file.', level=messages.ERROR)
                return HttpResponseRedirect(request.path)

            if not file_obj.name.lower().endswith('.xlsx'):
                self.message_user(request, 'Only .xlsx files are supported.', level=messages.ERROR)
                return HttpResponseRedirect(request.path)

            try:
                workbook = load_workbook(file_obj, data_only=True)
                sheet = workbook.active
            except Exception as exc:
                self.message_user(request, f'Failed to read file: {exc}', level=messages.ERROR)
                return HttpResponseRedirect(request.path)

            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                self.message_user(request, 'The file is empty.', level=messages.ERROR)
                return HttpResponseRedirect(request.path)

            headers = [self._normalize_header(h) for h in rows[0]]
            index_by_name = {name: i for i, name in enumerate(headers) if name}

            title_idx = index_by_name.get('title')
            author_idx = index_by_name.get('auth')
            if author_idx is None:
                author_idx = index_by_name.get('author')
            isbn_idx = index_by_name.get('isbn')
            barcode_idx = index_by_name.get('barcode')
            quantity_idx = index_by_name.get('quantiy')
            if quantity_idx is None:
                quantity_idx = index_by_name.get('quantity')

            if title_idx is None or author_idx is None:
                self.message_user(
                    request,
                    'Required columns missing. Expected at least: title, auth.',
                    level=messages.ERROR,
                )
                return HttpResponseRedirect(request.path)

            created_count = 0
            updated_count = 0
            skipped_count = 0

            for row in rows[1:]:
                title = self._to_text(row[title_idx] if title_idx < len(row) else None)
                author = self._to_text(row[author_idx] if author_idx < len(row) else None)
                isbn = self._to_text(row[isbn_idx] if isbn_idx is not None and isbn_idx < len(row) else None)
                barcode = self._to_text(
                    row[barcode_idx] if barcode_idx is not None and barcode_idx < len(row) else None
                )
                quantity = self._to_int(
                    row[quantity_idx] if quantity_idx is not None and quantity_idx < len(row) else None,
                    default=1,
                )

                if not title or not author:
                    skipped_count += 1
                    continue

                lookup = None
                if isbn:
                    lookup = Book.objects.filter(isbn=isbn).first()
                elif barcode:
                    lookup = Book.objects.filter(barcode=barcode).first()
                if lookup is None:
                    lookup = Book.objects.filter(title=title, author=author).first()

                payload = {
                    'title': title,
                    'author': author,
                    'isbn': isbn,
                    'barcode': barcode,
                    'quantity': quantity,
                    'available': quantity > 0,
                }

                if lookup:
                    for key, value in payload.items():
                        setattr(lookup, key, value)
                    lookup.save()
                    updated_count += 1
                else:
                    Book.objects.create(**payload)
                    created_count += 1

            self.message_user(
                request,
                f'Import finished: created={created_count}, updated={updated_count}, skipped={skipped_count}.',
                level=messages.SUCCESS,
            )
            return HttpResponseRedirect(reverse('admin:books_book_changelist'))

        context = dict(
            self.admin_site.each_context(request),
            opts=self.model._meta,
            title='Import books from XLSX',
        )
        return render(request, 'admin/books/book/import_xlsx.html', context)


@admin.register(Review)
class ReviewAdmin(admin.ModelAdmin):
    list_display = ['book', 'user', 'rating', 'created_at']
    list_filter = ['rating']


@admin.register(Favorite)
class FavoriteAdmin(admin.ModelAdmin):
    list_display = ['user', 'book', 'created_at']


@admin.register(BorrowRecord)
class BorrowRecordAdmin(admin.ModelAdmin):
    list_display = ['book', 'student', 'checked_out_at', 'checked_in_at', 'assigned_by', 'returned_by']
    list_filter = ['checked_in_at', 'checked_out_at']
    search_fields = ['book__title', 'book__barcode', 'student__username', 'student__student_profile__id_number']


@admin.register(Exhibit)
class ExhibitAdmin(admin.ModelAdmin):
    list_display = ['title', 'subtitle', 'date_start', 'date_end', 'is_active', 'order']
    list_editable = ['is_active', 'order']
    list_filter = ['is_active']
    search_fields = ['title']


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    list_display = ['title', 'category', 'date', 'time_start', 'time_end', 'location', 'is_active']
    list_editable = ['is_active']
    list_filter = ['is_active', 'date']
    search_fields = ['title', 'category', 'location']


@admin.register(News)
class NewsAdmin(admin.ModelAdmin):
    list_display = ['title', 'date', 'is_active']
    list_editable = ['is_active']
    list_filter = ['is_active', 'date']
    search_fields = ['title']
